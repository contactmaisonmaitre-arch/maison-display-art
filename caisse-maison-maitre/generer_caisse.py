#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de caisse — Maison Maitre
Usage :  python3 generer_caisse.py caisse_input.json
Sort un fichier Excel identique à la fiche papier (Billets / Pièces / Chèques / A->H),
avec formules vivantes. Nécessite : pip3 install openpyxl
"""
import sys, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build(data, out_path):
    F = "Arial"
    thin = Side(style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr = PatternFill("solid", fgColor="D9D9D9")
    sec = PatternFill("solid", fgColor="BFBFBF")
    res = PatternFill("solid", fgColor="E2EFDA")

    wb = Workbook(); ws = wb.active
    ws.title = "Caisse " + str(data.get("date", "")).replace("/", "-")[:28] or "Caisse"

    def C(ref, val=None, bold=False, size=11, fill=None, align="center",
          money=False, color="000000", wrap=False):
        c = ws[ref]
        if val is not None: c.value = val
        c.font = Font(name=F, bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = box
        if fill: c.fill = fill
        if money: c.number_format = '#,##0.00 "€";[Red]-#,##0.00 "€"'
        return c

    for col, w in {"A": 4, "B": 30, "C": 12, "D": 14, "E": 36}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:D1")
    C("A1", f"CAISSE  —  Date : {data.get('date','__________')}", bold=True, size=14, fill=hdr)
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:B2"); C("A2", "", fill=hdr)
    C("C2", "Nombre", bold=True, fill=hdr); C("D2", "Total", bold=True, fill=hdr)

    # --- BILLETS ---
    ws.merge_cells("A3:D3"); C("A3", "Billets", bold=True, fill=sec)
    order_b = ["100", "50", "20", "10", "5"]
    r = 4; b_first = r
    for lab in order_b:
        q = data.get("billets", {}).get(lab, 0)
        ws.merge_cells(f"A{r}:B{r}"); C(f"A{r}", lab, bold=True)
        C(f"C{r}", q, color="0000FF"); C(f"D{r}", f"=C{r}*{lab}", money=True)
        r += 1
    b_last = r - 1

    # --- PIECES ---
    ws.merge_cells(f"A{r}:D{r}"); C(f"A{r}", "Pièces", bold=True, fill=sec); r += 1
    order_p = ["2", "1", "0.5", "0.2", "0.1", "0.05", "0.02", "0.01"]
    labels = {"0.5": "0,5", "0.2": "0,2", "0.1": "0,1",
              "0.05": "0,05", "0.02": "0,02", "0.01": "0,01"}
    p_first = r
    for lab in order_p:
        q = data.get("pieces", {}).get(lab, 0)
        ws.merge_cells(f"A{r}:B{r}"); C(f"A{r}", labels.get(lab, lab), bold=True)
        C(f"C{r}", q, color="0000FF"); C(f"D{r}", f"=C{r}*{lab}", money=True)
        r += 1
    p_last = r - 1

    # --- Sous-total espèces ---
    especes_row = r
    ws.merge_cells(f"A{r}:C{r}")
    C(f"A{r}", "Sous-total ESPÈCES", bold=True, fill=hdr, align="right")
    C(f"D{r}", f"=SUM(D{b_first}:D{b_last})+SUM(D{p_first}:D{p_last})", bold=True, money=True, fill=hdr)
    r += 1

    # --- CHÈQUES ---
    ws.merge_cells(f"A{r}:D{r}"); C(f"A{r}", "Chèques  (noté à part — hors caisse physique)", bold=True, fill=sec, size=10); r += 1
    cheques = data.get("cheques", []) or []
    if not cheques:
        ws.merge_cells(f"A{r}:B{r}"); C(f"A{r}", "—", align="left")
        C(f"C{r}", ""); C(f"D{r}", 0, money=True, color="0000FF"); r += 1
    else:
        for ch in cheques:
            ws.merge_cells(f"A{r}:B{r}")
            C(f"A{r}", ch.get("label", "chèque"), align="left", size=10)
            C(f"C{r}", ""); C(f"D{r}", ch.get("montant", 0), money=True, color="0000FF"); r += 1

    # --- Bloc A -> H ---
    A  = data.get("A_a_transmettre", 0)
    B  = data.get("B_rouleaux", 0)
    E_D_prev = data.get("E_jour_precedent_D", 0)          # D du jour précédent
    depot_prev = data.get("depot_jour_precedent", 0)       # déposé la veille -> retiré de E
    E_eff = round(E_D_prev - depot_prev, 2)                # report réel dans le tiroir
    Eprime = data.get("E_prime_depose", 0)                 # déposé aujourd'hui
    G  = data.get("G_ticket_especes_net", 0)

    note_E = (f"{E_D_prev:.2f} - {depot_prev:.2f} déposés la veille"
              if depot_prev else f"= D du jour précédent")

    rows = [
        ("A",  "Espèces à transmettre à la banque", A, "in", ""),
        ("B",  "Rouleaux et Billets non ouverts",   B, "in", ""),
        ("C",  "Total du Jour",                      f"=D{especes_row}", "f", "Espèces comptées (chèque exclu)"),
        ("D",  "Total CAISSE (A+B+C)",               None, "f", ""),
        ("E",  "Jour Précédent (report réel)",       E_eff, "in", note_E),
        ("E'", "Déposé à la Banque",                 Eprime, "in", ""),
        ("F",  "Différence (D - E - E')",            None, "f", ""),
        ("G",  "Ticket caisse du jour (espèces net)",G, "in", "189,90 - rendu monnaie, etc."),
        ("H",  "F - G",                              None, "f", "Écart de caisse (doit être ~0)"),
    ]
    s = r
    rowref = {}
    for i, (lab, desc, val, kind, note) in enumerate(rows):
        rr = s + i; rowref[lab] = rr
        C(f"A{rr}", lab, bold=True, fill=hdr)
        C(f"B{rr}", desc, align="left", bold=(lab in ("C", "D", "F", "H")), wrap=True)
        C(f"C{rr}", "")
        if lab == "D":
            fa = f"=D{rowref['A']}+D{rowref['B']}+D{rowref['C']}"
            C(f"D{rr}", fa, bold=True, money=True)
        elif lab == "F":
            ep = rowref["E'"]
            ff = f"=D{rowref['D']}-D{rowref['E']}-D{ep}"
            C(f"D{rr}", ff, bold=True, money=True, fill=res)
        elif lab == "H":
            fh = f"=D{rowref['F']}-D{rowref['G']}"
            C(f"D{rr}", fh, bold=True, money=True, fill=res)
        elif kind == "in":
            C(f"D{rr}", val, money=True, color="0000FF", bold=(lab in ("C", "D", "F", "H")))
        else:
            C(f"D{rr}", val, money=True, bold=True)
        if note:
            ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=5)
            n = ws[f"E{rr}"]; n.value = note
            n.font = Font(name=F, italic=True, size=9, color="808080")
            n.alignment = Alignment(horizontal="left", vertical="center")

    lr = s + len(rows) + 1
    ws.merge_cells(f"A{lr}:E{lr}"); lc = ws[f"A{lr}"]
    lc.value = ("Bleu = saisi  •  Noir = calcul auto  •  E = report réel (dépôt de la veille déduit)  "
                "•  Chèque hors total caisse  •  G = espèces net du ticket")
    lc.font = Font(name=F, italic=True, size=9, color="404040")
    lc.alignment = Alignment(horizontal="left")

    wb.save(out_path)

    # Récap console
    especes = sum(int(data.get("billets", {}).get(l, 0)) * float(l) for l in order_b) \
            + sum(int(data.get("pieces", {}).get(l, 0)) * float(l) for l in order_p)
    D = round(A + B + especes, 2)
    Fv = round(D - E_eff - Eprime, 2)
    H = round(Fv - G, 2)
    print(f"Fichier : {out_path}")
    print(f"C (Total du Jour) = {especes:.2f} €")
    print(f"D (Total CAISSE)  = {D:.2f} €")
    print(f"E (report réel)   = {E_eff:.2f} €")
    print(f"F (Différence)    = {Fv:.2f} €")
    print(f"G (ticket)        = {G:.2f} €")
    print(f"H (F - G)         = {H:+.2f} €   {'OK ✅' if abs(H) < 2 else '⚠️ écart à vérifier'}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python3 generer_caisse.py caisse_input.json"); sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)
    out = data.get("fichier_sortie") or f"Caisse_{str(data.get('date','')).replace('/','-')}.xlsx"
    build(data, out)
